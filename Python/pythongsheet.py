import gspread
from openpyxl import load_workbook
from yaml import load, SafeLoader

# Read the filename from settings.yml
with open("settings.yaml","r",encoding="UTF-8") as stream:
    settings = load(stream,SafeLoader)
filename = settings['filename']
gs_name = settings['googlesheet_name']

# sa = gspread.service_account()
# sh = sa.open(gs_name)

# wks = sh.worksheet("main")
# print(wks.row_count)

book = load_workbook(filename)
print(book.sheetnames)
